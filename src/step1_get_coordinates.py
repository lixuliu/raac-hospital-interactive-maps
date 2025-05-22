"""
Aim: To retrieve geographic coordinates (latitude and longitude) for hospital postcodes using the Nominatim API and save these coordinates to an Excel file.

Objective: This script reads hospital postcodes from an Excel file, geocodes each postcode to obtain its latitude and longitude, and then appends these coordinates to the original Excel file in new columns.
"""

import openpyxl
import requests

def geocode_address(postcode):
    url = f'https://nominatim.openstreetmap.org/search?q={postcode}, UK&format=json&limit=1'
    response = requests.get(url)
    data = response.json()
    if data:
        location = data[0]
        return float(location['lat']), float(location['lon'])
    else:
        return None

def save_coordinates_to_xlsx(input_file, output_file):
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active

    # Add new columns for latitude and longitude
    sheet.cell(row=1, column=6).value = 'Latitude'
    sheet.cell(row=1, column=7).value = 'Longitude'

    for row in range(2, sheet.max_row + 1):
        postcode = sheet.cell(row=row, column=5).value
        if postcode:
            coordinates = geocode_address(postcode)
            if coordinates:
                sheet.cell(row=row, column=6).value = coordinates[0]
                sheet.cell(row=row, column=7).value = coordinates[1]
            else:
                sheet.cell(row=row, column=6).value = 'Coordinates not found'

    wb.save(output_file)
    print(f'Coordinates saved to {output_file}')

# Example file paths
input_file_path = "data/NGL15- Hospital_sites_with_confirmed_RAAC_list.xlsx"
output_file_path = "data/coordinates.xlsx"


save_coordinates_to_xlsx(input_file_path, output_file_path)
